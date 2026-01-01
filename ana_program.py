# ana_program.py

from basit_yardimcilar import *
from veri_islemleri import *

# Global değişkenler
VERI_DOSYASI = "hastane_depo.json"
sistem_verisi = None


def sistemi_baslat():
    """Programı başlat, verileri yükle"""
    global sistem_verisi

    baslik_yazdir("HASTANE DEPO YÖNETİM SİSTEMİ")

    sistem_verisi = veri_yukle(VERI_DOSYASI)

    if sistem_verisi is None:
        bilgi_mesaji("İlk çalıştırma. Yeni veri tabanı oluşturuluyor...")
        sistem_verisi = ilk_veri_olustur()
        veri_kaydet(VERI_DOSYASI, sistem_verisi)
        basarili_mesaji("Sistem hazır!")
    else:
        basarili_mesaji("Veriler yüklendi!")


def yeni_urun_id_olustur():
    """URN001, URN002... şeklinde ID oluştur"""
    mevcut_urunler = sistem_verisi["urunler"]

    if len(mevcut_urunler) == 0:
        return "URN001"

    en_buyuk_numara = 0
    for urun_id in mevcut_urunler.keys():
        numara = int(urun_id.replace("URN", ""))
        if numara > en_buyuk_numara:
            en_buyuk_numara = numara

    yeni_numara = en_buyuk_numara + 1
    return f"URN{yeni_numara:03d}"


def urun_ekle():
    """Yeni ürün ekle"""
    baslik_yazdir("YENİ ÜRÜN EKLE")

    urun_adi = input("Ürün adı: ")

    print("\nBarkod Tipi Seçin:")
    print("1. QR Kod")
    print("2. UTC Kod")
    barkod_secim = input("Seçiminiz (1/2): ")

    if barkod_secim == "1":
        barkod_tipi = "QR"
    elif barkod_secim == "2":
        barkod_tipi = "UTC"
    else:
        hata_mesaji("Geçersiz seçim!")
        return

    barkod_no = input("Barkod numarası: ")

    urun_id = yeni_urun_id_olustur()

    sistem_verisi["urunler"][urun_id] = {
        "ad": urun_adi,
        "barkod_tipi": barkod_tipi,
        "barkod_no": barkod_no
    }

    basarili_mesaji(f"Ürün eklendi! ID: {urun_id}")
    veri_kaydet(VERI_DOSYASI, sistem_verisi)


def urunleri_listele():
    """Tüm ürünleri göster"""
    baslik_yazdir("ÜRÜN LİSTESİ")

    urunler = sistem_verisi["urunler"]

    if len(urunler) == 0:
        bilgi_mesaji("Henüz ürün eklenmemiş.")
        return

    print(f"{'ID':<10} {'Ürün Adı':<30} {'Barkod Tipi':<15} {'Barkod No':<15}")
    print("-" * 70)

    for urun_id, urun_bilgi in urunler.items():
        print(f"{urun_id:<10} {urun_bilgi['ad']:<30} {urun_bilgi['barkod_tipi']:<15} {urun_bilgi['barkod_no']:<15}")


def urun_menu():
    """Ürün işlemleri menüsü"""
    while True:
        baslik_yazdir("ÜRÜN İŞLEMLERİ")
        print("1. Yeni Ürün Ekle")
        print("2. Ürünleri Listele")
        print("0. Ana Menüye Dön")
        print("-" * 50)

        secim = input("Seçiminiz: ")

        if secim == "1":
            urun_ekle()
        elif secim == "2":
            urunleri_listele()
            input("\nDevam etmek için Enter'a basın...")
        elif secim == "0":
            break
        else:
            hata_mesaji("Geçersiz seçim!")


# ===============================
# DEPO İŞLEMLERİ FONKSİYONLARI
# ===============================

def yeni_depo_id_olustur():
    """DEP001, DEP002... şeklinde ID oluştur"""
    mevcut_depolar = sistem_verisi["depolar"]

    if len(mevcut_depolar) == 0:
        return "DEP001"

    en_buyuk_numara = 0
    for depo_id in mevcut_depolar.keys():
        numara = int(depo_id.replace("DEP", ""))
        if numara > en_buyuk_numara:
            en_buyuk_numara = numara

    yeni_numara = en_buyuk_numara + 1
    return f"DEP{yeni_numara:03d}"


def depo_ekle():
    """Yeni depo ekle"""
    baslik_yazdir("YENİ DEPO EKLE")

    depo_adi = input("Depo adı (örn: Eczane Deposu): ")

    # Boş kontrol
    if depo_adi.strip() == "":
        hata_mesaji("Depo adı boş olamaz!")
        return

    depo_id = yeni_depo_id_olustur()

    # Depoyu sisteme ekle
    sistem_verisi["depolar"][depo_id] = {
        "ad": depo_adi,
        "urunler": {}  # Başlangıçta ürün yok
    }

    basarili_mesaji(f"Depo eklendi! ID: {depo_id}")
    veri_kaydet(VERI_DOSYASI, sistem_verisi)


def depolari_listele():
    """Tüm depoları göster"""
    baslik_yazdir("DEPO LİSTESİ")

    depolar = sistem_verisi["depolar"]

    if len(depolar) == 0:
        bilgi_mesaji("Henüz depo eklenmemiş.")
        return

    print(f"{'ID':<10} {'Depo Adı':<40} {'Ürün Sayısı':<15}")
    print("-" * 65)

    for depo_id, depo_bilgi in depolar.items():
        urun_sayisi = len(depo_bilgi["urunler"])
        print(f"{depo_id:<10} {depo_bilgi['ad']:<40} {urun_sayisi:<15}")


def depo_detay_goster():
    """Bir deponun detayını göster"""
    baslik_yazdir("DEPO DETAYI")

    depolar = sistem_verisi["depolar"]

    if len(depolar) == 0:
        bilgi_mesaji("Henüz depo eklenmemiş.")
        return

    # Önce depoları listele
    print("\nMevcut Depolar:")
    for depo_id, depo_bilgi in depolar.items():
        print(f"{depo_id} - {depo_bilgi['ad']}")

    print()
    depo_id = input("Depo ID'si girin: ").upper()

    # Depo var mı kontrol et
    if depo_id not in depolar:
        hata_mesaji("Geçersiz depo ID'si!")
        return

    depo_bilgi = depolar[depo_id]

    print("\n" + "=" * 70)
    print(f"DEPO ADI: {depo_bilgi['ad']}")
    print(f"DEPO ID: {depo_id}")
    print("=" * 70)

    # Bu depodaki ürünler
    if len(depo_bilgi["urunler"]) == 0:
        print("\nBu depoda henüz ürün yok.")
    else:
        print(f"\n{'Ürün ID':<10} {'Ürün Adı':<25} {'Miktar':<10} {'Min':<8} {'Max':<8} {'Kritik':<10}")
        print("-" * 70)

        for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
            urun_adi = sistem_verisi["urunler"][urun_id]["ad"]
            print(
                f"{urun_id:<10} {urun_adi:<25} {stok_bilgi['miktar']:<10} {stok_bilgi['min_seviye']:<8} {stok_bilgi['max_seviye']:<8} {stok_bilgi['kritik_seviye']:<10}")


def depo_sil():
    """Depo sil"""
    baslik_yazdir("DEPO SİL")

    depolar = sistem_verisi["depolar"]

    if len(depolar) == 0:
        bilgi_mesaji("Henüz depo eklenmemiş.")
        return

    # Depoları listele
    print("\nMevcut Depolar:")
    for depo_id, depo_bilgi in depolar.items():
        urun_sayisi = len(depo_bilgi["urunler"])
        print(f"{depo_id} - {depo_bilgi['ad']} ({urun_sayisi} ürün)")

    print()
    depo_id = input("Silinecek depo ID'si: ").upper()

    if depo_id not in depolar:
        hata_mesaji("Geçersiz depo ID'si!")
        return

    # Onay al
    depo_adi = depolar[depo_id]["ad"]
    urun_sayisi = len(depolar[depo_id]["urunler"])

    if urun_sayisi > 0:
        print(f"\n[!] UYARI: Bu depoda {urun_sayisi} adet ürün var!")

    onay = input(f"'{depo_adi}' deposunu silmek istediğinize emin misiniz? (e/h): ")

    if onay.lower() == 'e':
        del sistem_verisi["depolar"][depo_id]
        basarili_mesaji(f"'{depo_adi}' deposu silindi!")
        veri_kaydet(VERI_DOSYASI, sistem_verisi)
    else:
        bilgi_mesaji("İşlem iptal edildi.")


def depo_menu():
    """Depo işlemleri menüsü"""
    while True:
        baslik_yazdir("DEPO İŞLEMLERİ")
        print("1. Yeni Depo Ekle")
        print("2. Depoları Listele")
        print("3. Depo Detayı Görüntüle")
        print("4. Depo Sil")
        print("0. Ana Menüye Dön")
        print("-" * 50)

        secim = input("Seçiminiz: ")

        if secim == "1":
            depo_ekle()
        elif secim == "2":
            depolari_listele()
            input("\nDevam etmek için Enter'a basın...")
        elif secim == "3":
            depo_detay_goster()
            input("\nDevam etmek için Enter'a basın...")
        elif secim == "4":
            depo_sil()
        elif secim == "0":
            break
        else:
            hata_mesaji("Geçersiz seçim!")


# ===============================
# STOK İŞLEMLERİ - YARDIMCI FONKSİYONLAR
# ===============================

def depo_sec():
    """Kullanıcıya depo seçtir, seçilen depo ID'sini döndür"""
    depolar = sistem_verisi["depolar"]

    if len(depolar) == 0:
        hata_mesaji("Henüz depo eklenmemiş!")
        return None

    print("\nMevcut Depolar:")
    for depo_id, depo_bilgi in depolar.items():
        print(f"{depo_id} - {depo_bilgi['ad']}")

    print()
    depo_id = input("Depo ID'si girin: ").upper()

    if depo_id not in depolar:
        hata_mesaji("Geçersiz depo ID'si!")
        return None

    return depo_id


def urun_sec_atanmamis(depo_id):
    """Depoda OLMAYAN ürünleri göster ve seçtir"""
    tum_urunler = sistem_verisi["urunler"]
    depo_urunleri = sistem_verisi["depolar"][depo_id]["urunler"]

    # Bu depoda olmayan ürünleri bul
    atanmamis_urunler = {}
    for urun_id, urun_bilgi in tum_urunler.items():
        if urun_id not in depo_urunleri:
            atanmamis_urunler[urun_id] = urun_bilgi

    if len(atanmamis_urunler) == 0:
        bilgi_mesaji("Bu depoya atanabilecek ürün kalmadı!")
        return None

    print("\nBu depoda OLMAYAN ürünler:")
    for urun_id, urun_bilgi in atanmamis_urunler.items():
        print(f"{urun_id} - {urun_bilgi['ad']}")

    print()
    urun_id = input("Ürün ID'si girin: ").upper()

    if urun_id not in atanmamis_urunler:
        hata_mesaji("Geçersiz ürün ID'si!")
        return None

    return urun_id


def urun_sec_atanmis(depo_id):
    """Depoda OLAN ürünleri göster ve seçtir"""
    depo_urunleri = sistem_verisi["depolar"][depo_id]["urunler"]

    if len(depo_urunleri) == 0:
        bilgi_mesaji("Bu depoda henüz ürün yok!")
        return None

    print("\nBu depodaki ürünler:")
    for urun_id, stok_bilgi in depo_urunleri.items():
        urun_adi = sistem_verisi["urunler"][urun_id]["ad"]
        print(f"{urun_id} - {urun_adi} (Mevcut: {stok_bilgi['miktar']})")

    print()
    urun_id = input("Ürün ID'si girin: ").upper()

    if urun_id not in depo_urunleri:
        hata_mesaji("Geçersiz ürün ID'si!")
        return None

    return urun_id


def seviye_kontrol(kritik, min_seviye, max_seviye):
    """Seviyelerin mantıklı olup olmadığını kontrol et"""
    if kritik < 0 or min_seviye < 0 or max_seviye < 0:
        hata_mesaji("Seviyeler negatif olamaz!")
        return False

    if not (kritik < min_seviye < max_seviye):
        hata_mesaji("Mantıklı sıralama: Kritik < Minimum < Maksimum olmalı!")
        return False

    return True


def stok_durumu_hesapla(miktar, kritik, min_seviye, max_seviye):
    """Stok durumunu hesapla ve emoji ile döndür"""
    if miktar <= kritik:
        return "🔴 KRİTİK"
    elif miktar < min_seviye:
        return "🟡 Düşük"
    elif miktar <= max_seviye:
        return "🟢 Normal"
    else:
        return "🟠 Fazla"


def sayi_al(mesaj, minimum=0):
    """Kullanıcıdan geçerli bir sayı al"""
    while True:
        try:
            deger = int(input(mesaj))
            if deger < minimum:
                hata_mesaji(f"Değer en az {minimum} olmalı!")
                continue
            return deger
        except ValueError:
            hata_mesaji("Lütfen geçerli bir sayı girin!")


# ===============================
# STOK İŞLEMLERİ - ANA FONKSİYONLAR
# ===============================

def depoya_urun_ata():
    """Bir depoya ilk kez ürün ata"""
    baslik_yazdir("DEPOYA ÜRÜN ATA")

    # 1. Depo seç
    depo_id = depo_sec()
    if depo_id is None:
        return

    depo_adi = sistem_verisi["depolar"][depo_id]["ad"]
    print(f"\nSeçilen Depo: {depo_adi}")

    # 2. Ürün seç (bu depoda olmayan)
    urun_id = urun_sec_atanmamis(depo_id)
    if urun_id is None:
        return

    urun_adi = sistem_verisi["urunler"][urun_id]["ad"]
    print(f"Seçilen Ürün: {urun_adi}")

    # 3. Bilgileri al
    print("\n--- Stok Bilgileri ---")
    baslangic_miktar = sayi_al("Başlangıç miktarı: ", 0)
    kritik_seviye = sayi_al("Kritik seviye: ", 0)
    min_seviye = sayi_al("Minimum seviye: ", 0)
    max_seviye = sayi_al("Maksimum seviye: ", 0)

    # 4. Seviye kontrolü
    if not seviye_kontrol(kritik_seviye, min_seviye, max_seviye):
        return

    # Başlangıç miktarı kontrolü
    if baslangic_miktar > max_seviye:
        bilgi_mesaji(f"UYARI: Başlangıç miktarı ({baslangic_miktar}) maksimum seviyeyi ({max_seviye}) aşıyor!")
        onay = input("Yine de devam edilsin mi? (e/h): ")
        if onay.lower() != 'e':
            bilgi_mesaji("İşlem iptal edildi.")
            return

    # 5. Ürünü depoya ata
    sistem_verisi["depolar"][depo_id]["urunler"][urun_id] = {
        "miktar": baslangic_miktar,
        "kritik_seviye": kritik_seviye,
        "min_seviye": min_seviye,
        "max_seviye": max_seviye
    }

    basarili_mesaji(f"'{urun_adi}' başarıyla '{depo_adi}' deposuna atandı!")
    veri_kaydet(VERI_DOSYASI, sistem_verisi)

    # Durum göster
    durum = stok_durumu_hesapla(baslangic_miktar, kritik_seviye, min_seviye, max_seviye)
    print(f"Başlangıç Durumu: {durum}")


def stok_giris_yap():
    """Mevcut stoğa giriş yap (artır)"""
    baslik_yazdir("STOK GİRİŞ")

    # 1. Depo seç
    depo_id = depo_sec()
    if depo_id is None:
        return

    depo_adi = sistem_verisi["depolar"][depo_id]["ad"]
    print(f"\nSeçilen Depo: {depo_adi}")

    # 2. Ürün seç (bu depoda olan)
    urun_id = urun_sec_atanmis(depo_id)
    if urun_id is None:
        return

    urun_adi = sistem_verisi["urunler"][urun_id]["ad"]
    stok_bilgi = sistem_verisi["depolar"][depo_id]["urunler"][urun_id]

    # Mevcut durumu göster
    print(f"\n--- Mevcut Durum ---")
    print(f"Ürün: {urun_adi}")
    print(f"Mevcut Miktar: {stok_bilgi['miktar']}")
    print(f"Maksimum Seviye: {stok_bilgi['max_seviye']}")

    # 3. Giriş miktarı al
    giris_miktari = sayi_al("\nGiriş miktarı: ", 1)

    # 4. Yeni miktar hesapla
    yeni_miktar = stok_bilgi['miktar'] + giris_miktari

    print(f"\nYeni miktar: {stok_bilgi['miktar']} + {giris_miktari} = {yeni_miktar}")

    # 5. Max kontrolü
    if yeni_miktar > stok_bilgi['max_seviye']:
        bilgi_mesaji(f"[!] UYARI: Maksimum seviye ({stok_bilgi['max_seviye']}) aşılacak!")
        onay = input("Yine de devam edilsin mi? (e/h): ")
        if onay.lower() != 'e':
            bilgi_mesaji("İşlem iptal edildi.")
            return

    # 6. Güncelle ve kaydet
    sistem_verisi["depolar"][depo_id]["urunler"][urun_id]["miktar"] = yeni_miktar
    veri_kaydet(VERI_DOSYASI, sistem_verisi)

    basarili_mesaji(f"Stok girişi tamamlandı! Yeni miktar: {yeni_miktar}")

    # Durum göster
    durum = stok_durumu_hesapla(
        yeni_miktar,
        stok_bilgi['kritik_seviye'],
        stok_bilgi['min_seviye'],
        stok_bilgi['max_seviye']
    )
    print(f"Durum: {durum}")


def stok_cikis_yap():
    """Mevcut stoktan çıkış yap (azalt)"""
    baslik_yazdir("STOK ÇIKIŞ")

    # 1. Depo seç
    depo_id = depo_sec()
    if depo_id is None:
        return

    depo_adi = sistem_verisi["depolar"][depo_id]["ad"]
    print(f"\nSeçilen Depo: {depo_adi}")

    # 2. Ürün seç
    urun_id = urun_sec_atanmis(depo_id)
    if urun_id is None:
        return

    urun_adi = sistem_verisi["urunler"][urun_id]["ad"]
    stok_bilgi = sistem_verisi["depolar"][depo_id]["urunler"][urun_id]

    # Mevcut durumu göster
    print(f"\n--- Mevcut Durum ---")
    print(f"Ürün: {urun_adi}")
    print(f"Mevcut Miktar: {stok_bilgi['miktar']}")
    print(f"Minimum Seviye: {stok_bilgi['min_seviye']}")
    print(f"Kritik Seviye: {stok_bilgi['kritik_seviye']}")

    # 3. Çıkış miktarı al
    cikis_miktari = sayi_al("\nÇıkış miktarı: ", 1)

    # 4. Yeterli stok kontrolü
    if cikis_miktari > stok_bilgi['miktar']:
        hata_mesaji(f"Yetersiz stok! Mevcut: {stok_bilgi['miktar']}, İstenen: {cikis_miktari}")
        bilgi_mesaji("İşlem iptal edildi.")
        return

    # 5. Yeni miktar hesapla
    yeni_miktar = stok_bilgi['miktar'] - cikis_miktari

    print(f"\nYeni miktar: {stok_bilgi['miktar']} - {cikis_miktari} = {yeni_miktar}")

    # 6. Seviye kontrolleri
    uyari_var = False

    if yeni_miktar <= stok_bilgi['kritik_seviye']:
        hata_mesaji(f"[🔴 KRİTİK] Stok kritik seviyeye ({stok_bilgi['kritik_seviye']}) düşecek veya altına inecek!")
        uyari_var = True
    elif yeni_miktar < stok_bilgi['min_seviye']:
        bilgi_mesaji(f"[🟡 UYARI] Stok minimum seviyenin ({stok_bilgi['min_seviye']}) altına düşecek!")
        uyari_var = True

    if uyari_var:
        onay = input("Yine de devam edilsin mi? (e/h): ")
        if onay.lower() != 'e':
            bilgi_mesaji("İşlem iptal edildi.")
            return

    # 7. Güncelle ve kaydet
    sistem_verisi["depolar"][depo_id]["urunler"][urun_id]["miktar"] = yeni_miktar
    veri_kaydet(VERI_DOSYASI, sistem_verisi)

    basarili_mesaji(f"Stok çıkışı tamamlandı! Yeni miktar: {yeni_miktar}")

    # Durum göster
    durum = stok_durumu_hesapla(
        yeni_miktar,
        stok_bilgi['kritik_seviye'],
        stok_bilgi['min_seviye'],
        stok_bilgi['max_seviye']
    )
    print(f"Durum: {durum}")


def seviye_ayarla():
    """Ürünün min-max-kritik seviyelerini güncelle"""
    baslik_yazdir("SEVİYE AYARLAMA")

    # 1. Depo seç
    depo_id = depo_sec()
    if depo_id is None:
        return

    depo_adi = sistem_verisi["depolar"][depo_id]["ad"]
    print(f"\nSeçilen Depo: {depo_adi}")

    # 2. Ürün seç
    urun_id = urun_sec_atanmis(depo_id)
    if urun_id is None:
        return

    urun_adi = sistem_verisi["urunler"][urun_id]["ad"]
    stok_bilgi = sistem_verisi["depolar"][depo_id]["urunler"][urun_id]

    # Mevcut değerleri göster
    print(f"\n--- Mevcut Değerler ---")
    print(f"Ürün: {urun_adi}")
    print(f"Mevcut Miktar: {stok_bilgi['miktar']}")
    print(f"Kritik Seviye: {stok_bilgi['kritik_seviye']}")
    print(f"Minimum Seviye: {stok_bilgi['min_seviye']}")
    print(f"Maksimum Seviye: {stok_bilgi['max_seviye']}")

    # 3. Yeni değerleri al
    print("\n--- Yeni Değerler (boş bırakılırsa değişmez) ---")

    yeni_kritik = input(f"Yeni kritik seviye [{stok_bilgi['kritik_seviye']}]: ")
    yeni_min = input(f"Yeni minimum seviye [{stok_bilgi['min_seviye']}]: ")
    yeni_max = input(f"Yeni maksimum seviye [{stok_bilgi['max_seviye']}]: ")

    # Boş bırakılanları eski değerle doldur
    try:
        kritik = int(yeni_kritik) if yeni_kritik.strip() != "" else stok_bilgi['kritik_seviye']
        min_sev = int(yeni_min) if yeni_min.strip() != "" else stok_bilgi['min_seviye']
        max_sev = int(yeni_max) if yeni_max.strip() != "" else stok_bilgi['max_seviye']
    except ValueError:
        hata_mesaji("Geçersiz sayı girişi!")
        return

    # 4. Kontrol et
    if not seviye_kontrol(kritik, min_sev, max_sev):
        return

    # 5. Güncelle ve kaydet
    sistem_verisi["depolar"][depo_id]["urunler"][urun_id]["kritik_seviye"] = kritik
    sistem_verisi["depolar"][depo_id]["urunler"][urun_id]["min_seviye"] = min_sev
    sistem_verisi["depolar"][depo_id]["urunler"][urun_id]["max_seviye"] = max_sev

    veri_kaydet(VERI_DOSYASI, sistem_verisi)

    basarili_mesaji("Seviyeler başarıyla güncellendi!")

    # Yeni durumu göster
    durum = stok_durumu_hesapla(stok_bilgi['miktar'], kritik, min_sev, max_sev)
    print(f"Güncel Durum: {durum}")


def stok_raporu_goster():
    """Bir deponun stok durumunu detaylı göster"""
    baslik_yazdir("STOK RAPORU")

    # Depo seç
    depo_id = depo_sec()
    if depo_id is None:
        return

    depo_bilgi = sistem_verisi["depolar"][depo_id]

    print("\n" + "=" * 90)
    print(f"{depo_bilgi['ad'].upper()} - STOK DURUMU".center(90))
    print("=" * 90)

    if len(depo_bilgi["urunler"]) == 0:
        bilgi_mesaji("Bu depoda henüz ürün yok.")
        return

    # Başlık
    print(f"{'Ürün ID':<10} {'Ürün Adı':<25} {'Miktar':<10} {'Kritik':<10} {'Min':<10} {'Max':<10} {'Durum':<15}")
    print("-" * 90)

    # Her ürünü listele
    for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
        urun_adi = sistem_verisi["urunler"][urun_id]["ad"]
        durum = stok_durumu_hesapla(
            stok_bilgi['miktar'],
            stok_bilgi['kritik_seviye'],
            stok_bilgi['min_seviye'],
            stok_bilgi['max_seviye']
        )

        print(
            f"{urun_id:<10} {urun_adi:<25} {stok_bilgi['miktar']:<10} {stok_bilgi['kritik_seviye']:<10} {stok_bilgi['min_seviye']:<10} {stok_bilgi['max_seviye']:<10} {durum:<15}")

    print("=" * 90)


def stok_menu():
    """Stok işlemleri menüsü"""
    while True:
        baslik_yazdir("STOK İŞLEMLERİ")
        print("1. Depoya Ürün Ata (İlk Kez)")
        print("2. Stok Giriş Yap")
        print("3. Stok Çıkış Yap")
        print("4. Min-Max-Kritik Ayarla")
        print("5. Stok Raporu Görüntüle")
        print("0. Ana Menüye Dön")
        print("-" * 50)

        secim = input("Seçiminiz: ")

        if secim == "1":
            depoya_urun_ata()
        elif secim == "2":
            stok_giris_yap()
        elif secim == "3":
            stok_cikis_yap()
        elif secim == "4":
            seviye_ayarla()
        elif secim == "5":
            stok_raporu_goster()
            input("\nDevam etmek için Enter'a basın...")
        elif secim == "0":
            break
        else:
            hata_mesaji("Geçersiz seçim!")


# ===============================
# CHATBOX - YARDIMCI FONKSİYONLAR
# ===============================

import csv
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def tarih_saat_al():
    """Şu anki tarih ve saati formatla"""
    return datetime.now().strftime("%d.%m.%Y %H:%M")


def kritik_urunleri_bul():
    """Tüm depolardan kritik seviyedeki ürünleri bul"""
    sonuclar = []

    for depo_id, depo_bilgi in sistem_verisi["depolar"].items():
        for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
            if stok_bilgi["miktar"] <= stok_bilgi["kritik_seviye"]:
                sonuclar.append({
                    "depo_id": depo_id,
                    "depo_adi": depo_bilgi["ad"],
                    "urun_id": urun_id,
                    "urun_adi": sistem_verisi["urunler"][urun_id]["ad"],
                    "miktar": stok_bilgi["miktar"],
                    "kritik": stok_bilgi["kritik_seviye"],
                    "min": stok_bilgi["min_seviye"],
                    "max": stok_bilgi["max_seviye"],
                    "durum": "🔴 Kritik"
                })

    return sonuclar


def dusuk_urunleri_bul():
    """Minimum seviyenin altındaki ürünleri bul"""
    sonuclar = []

    for depo_id, depo_bilgi in sistem_verisi["depolar"].items():
        for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
            if stok_bilgi["kritik_seviye"] < stok_bilgi["miktar"] < stok_bilgi["min_seviye"]:
                sonuclar.append({
                    "depo_id": depo_id,
                    "depo_adi": depo_bilgi["ad"],
                    "urun_id": urun_id,
                    "urun_adi": sistem_verisi["urunler"][urun_id]["ad"],
                    "miktar": stok_bilgi["miktar"],
                    "kritik": stok_bilgi["kritik_seviye"],
                    "min": stok_bilgi["min_seviye"],
                    "max": stok_bilgi["max_seviye"],
                    "durum": "🟡 Düşük"
                })

    return sonuclar


def normal_urunleri_bul():
    """Normal seviyedeki ürünleri bul"""
    sonuclar = []

    for depo_id, depo_bilgi in sistem_verisi["depolar"].items():
        for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
            if stok_bilgi["min_seviye"] <= stok_bilgi["miktar"] <= stok_bilgi["max_seviye"]:
                sonuclar.append({
                    "depo_id": depo_id,
                    "depo_adi": depo_bilgi["ad"],
                    "urun_id": urun_id,
                    "urun_adi": sistem_verisi["urunler"][urun_id]["ad"],
                    "miktar": stok_bilgi["miktar"],
                    "kritik": stok_bilgi["kritik_seviye"],
                    "min": stok_bilgi["min_seviye"],
                    "max": stok_bilgi["max_seviye"],
                    "durum": "🟢 Normal"
                })

    return sonuclar


def fazla_urunleri_bul():
    """Maksimum seviyenin üstündeki ürünleri bul"""
    sonuclar = []

    for depo_id, depo_bilgi in sistem_verisi["depolar"].items():
        for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
            if stok_bilgi["miktar"] > stok_bilgi["max_seviye"]:
                sonuclar.append({
                    "depo_id": depo_id,
                    "depo_adi": depo_bilgi["ad"],
                    "urun_id": urun_id,
                    "urun_adi": sistem_verisi["urunler"][urun_id]["ad"],
                    "miktar": stok_bilgi["miktar"],
                    "kritik": stok_bilgi["kritik_seviye"],
                    "min": stok_bilgi["min_seviye"],
                    "max": stok_bilgi["max_seviye"],
                    "durum": "🟠 Fazla"
                })

    return sonuclar


def kritige_yakin_urunleri_bul():
    """Kritik seviyeye yakın ürünleri bul (kritik + 5 aralığında)"""
    sonuclar = []

    for depo_id, depo_bilgi in sistem_verisi["depolar"].items():
        for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
            esik = stok_bilgi["kritik_seviye"] + 5
            if stok_bilgi["kritik_seviye"] < stok_bilgi["miktar"] <= esik:
                sonuclar.append({
                    "depo_id": depo_id,
                    "depo_adi": depo_bilgi["ad"],
                    "urun_id": urun_id,
                    "urun_adi": sistem_verisi["urunler"][urun_id]["ad"],
                    "miktar": stok_bilgi["miktar"],
                    "kritik": stok_bilgi["kritik_seviye"],
                    "min": stok_bilgi["min_seviye"],
                    "max": stok_bilgi["max_seviye"],
                    "durum": "⚠️ Kritik Yakın"
                })

    return sonuclar


def depo_urunlerini_bul(depo_adi_parcasi):
    """Belirli bir depodaki tüm ürünleri bul"""
    sonuclar = []

    for depo_id, depo_bilgi in sistem_verisi["depolar"].items():
        # Depo adında aranan kelime var mı?
        if depo_adi_parcasi.lower() in depo_bilgi["ad"].lower():
            for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
                durum = stok_durumu_hesapla(
                    stok_bilgi["miktar"],
                    stok_bilgi["kritik_seviye"],
                    stok_bilgi["min_seviye"],
                    stok_bilgi["max_seviye"]
                )

                sonuclar.append({
                    "depo_id": depo_id,
                    "depo_adi": depo_bilgi["ad"],
                    "urun_id": urun_id,
                    "urun_adi": sistem_verisi["urunler"][urun_id]["ad"],
                    "miktar": stok_bilgi["miktar"],
                    "kritik": stok_bilgi["kritik_seviye"],
                    "min": stok_bilgi["min_seviye"],
                    "max": stok_bilgi["max_seviye"],
                    "durum": durum
                })

    return sonuclar


def tum_depo_durumu():
    """Tüm depoların tüm ürünlerini getir"""
    sonuclar = []

    for depo_id, depo_bilgi in sistem_verisi["depolar"].items():
        for urun_id, stok_bilgi in depo_bilgi["urunler"].items():
            durum = stok_durumu_hesapla(
                stok_bilgi["miktar"],
                stok_bilgi["kritik_seviye"],
                stok_bilgi["min_seviye"],
                stok_bilgi["max_seviye"]
            )

            sonuclar.append({
                "depo_id": depo_id,
                "depo_adi": depo_bilgi["ad"],
                "urun_id": urun_id,
                "urun_adi": sistem_verisi["urunler"][urun_id]["ad"],
                "miktar": stok_bilgi["miktar"],
                "kritik": stok_bilgi["kritik_seviye"],
                "min": stok_bilgi["min_seviye"],
                "max": stok_bilgi["max_seviye"],
                "durum": durum
            })

    return sonuclar


# ===============================
# CHATBOX - FORMAT OLUŞTURMA
# ===============================

def ekranda_goster(sonuclar, baslik):
    """Sonuçları ekranda tablo olarak göster"""
    print("\n" + "=" * 100)
    print(baslik.upper().center(100))
    print(f"Tarih: {tarih_saat_al()}".center(100))
    print("=" * 100)

    if len(sonuclar) == 0:
        bilgi_mesaji("Sonuç bulunamadı.")
        return

    print(f"{'Depo':<30} {'Ürün':<25} {'Miktar':<10} {'Kritik':<10} {'Min':<10} {'Max':<10} {'Durum':<15}")
    print("-" * 100)

    for item in sonuclar:
        print(
            f"{item['depo_adi']:<30} {item['urun_adi']:<25} {item['miktar']:<10} {item['kritik']:<10} {item['min']:<10} {item['max']:<10} {item['durum']:<15}")

    print("=" * 100)
    print(f"Toplam: {len(sonuclar)} ürün")


def pdf_olustur(sonuclar, baslik):
    """PDF raporu oluştur"""
    tarih_str = datetime.now().strftime("%d_%m_%Y_%H%M")
    dosya_adi = f"{baslik.replace(' ', '_').lower()}_{tarih_str}.pdf"
    dosya_yolu = f"/mnt/user-data/outputs/{dosya_adi}"

    # PDF oluştur
    doc = SimpleDocTemplate(dosya_yolu, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Başlık
    baslik_style = styles['Title']
    elements.append(Paragraph(baslik.upper(), baslik_style))
    elements.append(Spacer(1, 0.5 * cm))

    # Tarih
    tarih_style = styles['Normal']
    elements.append(Paragraph(f"Tarih: {tarih_saat_al()}", tarih_style))
    elements.append(Spacer(1, 0.5 * cm))

    # Tablo verisi
    data = [['Depo', 'Ürün', 'Miktar', 'Kritik', 'Min', 'Max', 'Durum']]

    for item in sonuclar:
        data.append([
            item['depo_adi'],
            item['urun_adi'],
            str(item['miktar']),
            str(item['kritik']),
            str(item['min']),
            str(item['max']),
            item['durum'].replace('🔴', 'KRİTİK').replace('🟡', 'DÜŞÜK').replace('🟢', 'NORMAL').replace('🟠',
                                                                                                      'FAZLA').replace(
                '⚠️', 'YAKIN')
        ])

    # Tablo oluştur
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(f"Toplam: {len(sonuclar)} ürün", tarih_style))

    doc.build(elements)

    basarili_mesaji(f"PDF oluşturuldu: {dosya_adi}")
    return dosya_yolu


def excel_olustur(sonuclar, baslik):
    """Excel raporu oluştur"""
    tarih_str = datetime.now().strftime("%d_%m_%Y_%H%M")
    dosya_adi = f"{baslik.replace(' ', '_').lower()}_{tarih_str}.xlsx"
    dosya_yolu = f"/mnt/user-data/outputs/{dosya_adi}"

    # Workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Raporu"

    # Başlık
    ws.merge_cells('A1:G1')
    baslik_cell = ws['A1']
    baslik_cell.value = baslik.upper()
    baslik_cell.font = Font(size=16, bold=True)
    baslik_cell.alignment = Alignment(horizontal='center')

    # Tarih
    ws.merge_cells('A2:G2')
    tarih_cell = ws['A2']
    tarih_cell.value = f"Tarih: {tarih_saat_al()}"
    tarih_cell.alignment = Alignment(horizontal='center')

    # Başlıklar
    headers = ['Depo', 'Ürün', 'Miktar', 'Kritik', 'Min', 'Max', 'Durum']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Veriler
    for row, item in enumerate(sonuclar, start=5):
        ws.cell(row=row, column=1).value = item['depo_adi']
        ws.cell(row=row, column=2).value = item['urun_adi']
        ws.cell(row=row, column=3).value = item['miktar']
        ws.cell(row=row, column=4).value = item['kritik']
        ws.cell(row=row, column=5).value = item['min']
        ws.cell(row=row, column=6).value = item['max']

        durum_cell = ws.cell(row=row, column=7)
        durum_cell.value = item['durum'].replace('🔴', 'KRİTİK').replace('🟡', 'DÜŞÜK').replace('🟢', 'NORMAL').replace(
            '🟠', 'FAZLA').replace('⚠️', 'YAKIN')

        # Renklendirme
        if '🔴' in item['durum']:
            durum_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            durum_cell.font = Font(color="FFFFFF", bold=True)
        elif '🟡' in item['durum']:
            durum_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif '🟢' in item['durum']:
            durum_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif '🟠' in item['durum']:
            durum_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    wb.save(dosya_yolu)
    basarili_mesaji(f"Excel oluşturuldu: {dosya_adi}")
    return dosya_yolu


def csv_olustur(sonuclar, baslik):
    """CSV raporu oluştur"""
    tarih_str = datetime.now().strftime("%d_%m_%Y_%H%M")
    dosya_adi = f"{baslik.replace(' ', '_').lower()}_{tarih_str}.csv"
    dosya_yolu = f"/mnt/user-data/outputs/{dosya_adi}"

    with open(dosya_yolu, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['Tarih', 'Depo', 'Ürün', 'Miktar', 'Kritik', 'Minimum', 'Maksimum', 'Durum']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for item in sonuclar:
            writer.writerow({
                'Tarih': tarih_saat_al(),
                'Depo': item['depo_adi'],
                'Ürün': item['urun_adi'],
                'Miktar': item['miktar'],
                'Kritik': item['kritik'],
                'Minimum': item['min'],
                'Maksimum': item['max'],
                'Durum': item['durum'].replace('🔴', 'KRİTİK').replace('🟡', 'DÜŞÜK').replace('🟢', 'NORMAL').replace('🟠',
                                                                                                                   'FAZLA').replace(
                    '⚠️', 'YAKIN')
            })

    basarili_mesaji(f"CSV oluşturuldu: {dosya_adi}")
    return dosya_yolu


def format_sec_ve_olustur(sonuclar, baslik):
    """Kullanıcıya format seçtir ve ilgili dosyayı oluştur"""
    if len(sonuclar) == 0:
        bilgi_mesaji("Gösterilecek sonuç bulunamadı.")
        return

    print(f"\n{len(sonuclar)} adet sonuç bulundu.")
    print("\nHangi formatta görmek istersiniz?")
    print("1. Ekranda Göster")
    print("2. PDF olarak İndir")
    print("3. Excel (XLSX) olarak İndir")
    print("4. CSV olarak İndir")
    print("0. İptal")

    secim = input("\nSeçiminiz: ")

    if secim == "1":
        ekranda_goster(sonuclar, baslik)
    elif secim == "2":
        dosya = pdf_olustur(sonuclar, baslik)
        bilgi_mesaji(f"Dosya konumu: {dosya}")
    elif secim == "3":
        dosya = excel_olustur(sonuclar, baslik)
        bilgi_mesaji(f"Dosya konumu: {dosya}")
    elif secim == "4":
        dosya = csv_olustur(sonuclar, baslik)
        bilgi_mesaji(f"Dosya konumu: {dosya}")
    elif secim == "0":
        bilgi_mesaji("İptal edildi.")
    else:
        hata_mesaji("Geçersiz seçim!")


# ===============================
# CHATBOX - ANA FONKSİYON
# ===============================

def chatbox_analiz(mesaj):
    """Kullanıcı mesajını analiz et ve uygun komutu çalıştır"""
    mesaj = mesaj.lower().strip()

    # Kritik seviye
    if ("kritik" in mesaj or "kırmızı" in mesaj) and ("göster" in mesaj or "listele" in mesaj or "bul" in mesaj):
        if "yakın" in mesaj or "yaklaş" in mesaj:
            sonuclar = kritige_yakin_urunleri_bul()
            format_sec_ve_olustur(sonuclar, "Kritik Seviyeye Yakın Ürünler")
        else:
            sonuclar = kritik_urunleri_bul()
            format_sec_ve_olustur(sonuclar, "Kritik Seviyedeki Ürünler")
        return True

    # Düşük stok
    if ("düşük" in mesaj or "az" in mesaj or "minimum" in mesaj or "min" in mesaj) and (
            "göster" in mesaj or "listele" in mesaj):
        sonuclar = dusuk_urunleri_bul()
        format_sec_ve_olustur(sonuclar, "Minimum Seviyenin Altındaki Ürünler")
        return True

    # Normal stok
    if "normal" in mesaj and ("göster" in mesaj or "listele" in mesaj):
        sonuclar = normal_urunleri_bul()
        format_sec_ve_olustur(sonuclar, "Normal Seviyedeki Ürünler")
        return True

    # Fazla stok
    if ("fazla" in mesaj or "çok" in mesaj or "maksimum" in mesaj or "max" in mesaj) and (
            "göster" in mesaj or "listele" in mesaj):
        sonuclar = fazla_urunleri_bul()
        format_sec_ve_olustur(sonuclar, "Maksimum Seviyenin Üstündeki Ürünler")
        return True

    # Tüm depolar
    if ("tüm" in mesaj or "hepsi" in mesaj or "bütün" in mesaj) and "depo" in mesaj:
        sonuclar = tum_depo_durumu()
        format_sec_ve_olustur(sonuclar, "Tüm Depolar - Genel Durum")
        return True

    # Belirli depo
    if "depo" in mesaj and ("göster" in mesaj or "durum" in mesaj):
        # Depo adını bulmaya çalış
        kelimeler = mesaj.split()
        for i, kelime in enumerate(kelimeler):
            if kelime == "depo" and i > 0:
                depo_adi = kelimeler[i - 1]
                sonuclar = depo_urunlerini_bul(depo_adi)
                if len(sonuclar) > 0:
                    format_sec_ve_olustur(sonuclar, f"{sonuclar[0]['depo_adi']} - Stok Durumu")
                else:
                    hata_mesaji(f"'{depo_adi}' içeren depo bulunamadı!")
                return True

    # Genel durum
    if "genel" in mesaj and "durum" in mesaj:
        sonuclar = tum_depo_durumu()
        format_sec_ve_olustur(sonuclar, "Genel Sistem Durumu")
        return True

    return False


def chatbox_menu():
    """Chatbox ana menüsü"""
    baslik_yazdir("CHATBOX - AKILLI SORGULAMA")

    print("\n📋 Örnek Komutlar:")
    print("  • 'kritik seviyedeki ürünleri göster'")
    print("  • 'düşük stokları listele'")
    print("  • 'normal ürünleri göster'")
    print("  • 'fazla stokları bul'")
    print("  • 'kritik seviyeye yakın ürünler'")
    print("  • 'eczane deposunun durumunu göster'")
    print("  • 'tüm depoların durumunu göster'")
    print("  • 'genel durum'")
    print("  • 'çıkış' (Chatbox'tan çık)")

    while True:
        print("\n" + "-" * 50)
        mesaj = input("💬 Komut: ").strip()

        if mesaj.lower() in ['çıkış', 'exit', 'quit', 'q']:
            bilgi_mesaji("Chatbox'tan çıkılıyor...")
            break

        if mesaj == "":
            continue

        # Mesajı analiz et
        anlasildi = chatbox_analiz(mesaj)

        if not anlasildi:
            hata_mesaji("Komut anlaşılamadı. Lütfen örnek komutlara bakın.")


def ana_menu():
    """Ana menüyü göster"""
    while True:
        baslik_yazdir("ANA MENÜ")
        print("1. Ürün İşlemleri")
        print("2. Depo İşlemleri")
        print("3. Stok İşlemleri")
        print("4. Chatbox (Akıllı Sorgulama)")  # ← Artık çalışıyor!
        print("0. Çıkış")
        print("-" * 50)

        secim = input("Seçiminiz: ")

        if secim == "1":
            urun_menu()
        elif secim == "2":
            depo_menu()
        elif secim == "3":
            stok_menu()
        elif secim == "4":
            chatbox_menu()  # ← Ekledik
        elif secim == "0":
            if veri_kaydet(VERI_DOSYASI, sistem_verisi):
                basarili_mesaji("Veriler kaydedildi. Görüşmek üzere!")
            break
        else:
            hata_mesaji("Geçersiz seçim!")


